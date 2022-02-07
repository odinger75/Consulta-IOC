#libreria para menejo de excel
import openpyxl

#libreria para manejo del navegador
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time

options = webdriver.ChromeOptions()
options.add_argument("start-maximized");
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")


#abrir navegador en la pagina de fortinet
#driver = webdriver.Chrome('C:\chromedriver\chromedriver.exe')
chrome_service = Service(executable_path='C:\chromedriver\chromedriver.exe')
driver = webdriver.Chrome(service=chrome_service, options=options)
driver.get("https://www.fortiguard.com/webfilter")
#time.sleep(5)
WebDriverWait(driver, 15)


book = openpyxl.load_workbook('categoria.xlsx')
hoja = book.active
Columna=hoja.max_column
fila=hoja.max_row
total=fila
print(fila)
#print(fila)
for i in range (1,fila+1):
	ioc=hoja.cell(row = i, column =2).value
	print(i," de ",total,"    ",ioc)
	driver.find_element(By.XPATH,"/html/body/div[2]/div/div/div[2]/div[2]/div[2]/section/div[1]/div/form/input").send_keys(ioc)
	driver.find_element(By.XPATH,"//*[@id='submit']").send_keys(Keys.ENTER)
	time.sleep(10)

#extraer categoria ala que pertenece el ioc en fortinet
	categoria = driver.find_element(By.XPATH,"//*[@id='two-column']/div[2]/div[2]/section/div[3]/div/div/h4").text                                                             
	driver.get("https://www.fortiguard.com/webfilter")
	time.sleep(10)
	print(categoria)
#guarda la categoria en el mismo archivo donde esta el ioc	
	hoja.cell(row=i, column=3).value = categoria
	book.save('categoria.xlsx')

driver.close()
print ('Se Termino....')
